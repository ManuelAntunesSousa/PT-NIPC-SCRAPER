import requests
from bs4 import BeautifulSoup
import openpyxl
import logging
import re



# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Function to retrieve data
def get_data_for_nif(nif, view_state, event_validation):
    url = 'http://www.sicae.pt/Consulta.aspx'

    payload = {
        '__EVENTTARGET': '',
        '__EVENTARGUMENT': '',
        '__VIEWSTATE': view_state,
        '__EVENTVALIDATION': event_validation,
        'ctl00$MainContent$ipNipc': nif,  # Updated key for NIF
        'ctl00$MainContent$btnPesquisa': 'Pesquisar'
    }

    try:
        response = requests.post(url, data=payload, timeout=20)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        logger.error(f"Failed to retrieve data for NIF {nif}. Error: {e}")
        return None

    soup = BeautifulSoup(response.text, 'html.parser')

    # Find the NIPC value from the input field
    nipc_element = soup.find('input', {'id': 'ctl00_MainContent_ipNipc'})
    nipc_value = nipc_element.get('value') if nipc_element else 'N/A'

    
    cae_principal_table = soup.find('table', {'id': 'ctl00_MainContent_ConsultaDataGrid'})
    cae_principal = ' '.join(cae_principal_table.stripped_strings) if cae_principal_table else 'N/A'
 
    pattern = re.compile(r'(\d+)\s+([^0-?]+)\s+(\d+)\s')

    # Extracted info
    extracted_info = []
    # Matches in the data
    matches = re.findall(pattern, cae_principal)

    # Extracted info
    for match in matches:
        nipc = match[0]
        denom_social_firma = match[1]
        cae_principal_value = match[2]
        #extra_caes = match[3]

  
        return {'NIPC': nif, 'NIPC_read': nipc_value, 'Denominação Social': denom_social_firma, 'CAE Principal': cae_principal_value} #, 'CAE Secundário':extra_caes}

def main():

    # Load Excel file
    try:
        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb.active
    except Exception as e:
        logger.error(f"Error loading Excel file: {e}")
        return


    try:
        initial_response = requests.get('http://www.sicae.pt/Consulta.aspx', timeout=10)
        initial_response.raise_for_status()
    except requests.exceptions.RequestException as e:
        logger.error(f"Failed to make initial request. Error: {e}")
        wb.close()
        return

    initial_soup = BeautifulSoup(initial_response.text, 'html.parser')
    initial_view_state = initial_soup.find('input', {'id': '__VIEWSTATE'}).get('value', '')
    initial_event_validation = initial_soup.find('input', {'id': '__EVENTVALIDATION'}).get('value', '')

    
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=2):
        nif = row[0]
        logger.info(f"Processing NIF: {nif}")

        # Perform the data retrieval
        data = get_data_for_nif(nif, initial_view_state, initial_event_validation)

        if data:
            # Update the Excel file with the retrieved data
            sheet.cell(row=row_index, column=3, value=data['NIPC'])
            sheet.cell(row=row_index, column=4, value=data['NIPC_read'])
            sheet.cell(row=row_index, column=5, value=data['Denominação Social'])
            sheet.cell(row=row_index, column=6, value=data['CAE Principal'])
            # sheet.cell(row=row_index, column=7, value=data['CAE Secundário'])

    # Excel file
    try:
        wb.save(excel_file_path)
        logger.info("Excel file saved successfully.")
    except Exception as e:
        logger.error(f"Error saving Excel file: {e}")
    finally:
      
        logger.info("Data retrieval and update completed.")
        wb.close()

if __name__ == "__main__":
    excel_file_path = r'C:\Users\Manu\Dropbox\Data Colab\SGS\NIPC Scraper\nifdemo.xlsx'
    main()
